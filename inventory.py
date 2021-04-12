'''
This code import a spreadsheet, calculates and returns in a dictionary the number of products by supplier,
calculates and returns in a dictionary the value of inventory by supplier, and returns the items under 10 units.
'''

import openpyxl

inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    #Calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_producs = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_producs + 1

    else:
        products_per_supplier[supplier_name] = 1

    #Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price

    else:
        total_value_per_supplier[supplier_name] = inventory * price

    #Calculation and logic for inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # Add value for total inventory price
    inventory_price.value = inventory * price

print(f'Number of products by supplier: \n {products_per_supplier}')
print(f'Total value by supplier is: \n {total_value_per_supplier}')
print(f'Inventory under 10 by product number: \n (product num : quantity) \n {products_under_10_inv}')

inv_file.save('inventory_with_total_value.xlsx')